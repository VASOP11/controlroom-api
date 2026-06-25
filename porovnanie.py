import csv, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

gt_file = 'ground_truth.csv'
bench_file = 'benchmark_v6.15_results.csv'

# Load GT
with open(gt_file, encoding='utf-8') as f:
    gt_rows = {r['url'].rstrip('/'): r for r in csv.DictReader(f)}

# Load benchmark
with open(bench_file, encoding='utf-8-sig') as f:
    bench_rows = {r['url'].rstrip('/'): r for r in csv.DictReader(f)}

wb = Workbook()

# ── SHEET 1: Len 23 webov s menom (GOLDEN) ─────────────────
ws1 = wb.active
ws1.title = "GOLDEN - meno+rola"

headers = ['Web', 'GT Meno', 'GT Rola', 'Nasli Osoby', 'Meno OK?', 'GT Telefon', 'Nasli Telefony', 'Tel OK?', 'GT Email', 'Nasli Emaily', 'Email OK?']
green = PatternFill('solid', fgColor='C6EFCE')
red = PatternFill('solid', fgColor='FFC7CE')
yellow = PatternFill('solid', fgColor='FFEB9C')
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin

row = 2
for url, gt in sorted(gt_rows.items(), key=lambda x: x[0]):
    if not gt.get('meno', '').strip():
        continue
    
    domain = url.replace('https://','').replace('http://','').replace('www.','').split('/')[0]
    b = bench_rows.get(url, bench_rows.get(url+'/', {}))
    
    found_osoby = b.get('found_osoby', '')
    meno_match = b.get('meno_match', '')
    phone_match = b.get('phone_match', '')
    email_match = b.get('email_match', '')
    
    vals = [
        domain,
        gt.get('meno', ''),
        gt.get('rola', ''),
        found_osoby[:120],
        meno_match,
        gt.get('telefon', ''),
        b.get('found_phones', ''),
        phone_match,
        gt.get('emails', ''),
        b.get('found_emails', '')[:80],
        email_match
    ]
    
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=row, column=col, value=v)
        c.border = thin
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if col in (5, 8, 11):
            if v == 'Y':
                c.fill = green
                c.font = Font(bold=True, color='006100')
            elif v == 'N':
                c.fill = red
                c.font = Font(bold=True, color='9C0006')
    row += 1

ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 55
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 22
ws1.column_dimensions['G'].width = 35
ws1.column_dimensions['H'].width = 10
ws1.column_dimensions['I'].width = 30
ws1.column_dimensions['J'].width = 35
ws1.column_dimensions['K'].width = 10

# ── SHEET 2: Vsetky 109 webov ───────────────────────────────
ws2 = wb.create_sheet("Vsetky 109 webov")

headers2 = ['Web', 'Status', 'GT Email', 'Nasli Emaily', 'Email OK?', 'GT Telefon', 'Nasli Telefony', 'Tel OK?', 'GT Meno', 'Nasli Osoby', 'Meno OK?']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin

row = 2
for url, gt in sorted(gt_rows.items(), key=lambda x: x[0]):
    domain = url.replace('https://','').replace('http://','').replace('www.','').split('/')[0]
    b = bench_rows.get(url, bench_rows.get(url+'/', {}))
    
    has_name = bool(gt.get('meno', '').strip())
    
    vals = [
        domain,
        b.get('status', 'N/A'),
        gt.get('emails', ''),
        b.get('found_emails', '')[:60],
        b.get('email_match', ''),
        gt.get('telefon', ''),
        b.get('found_phones', '')[:50],
        b.get('phone_match', ''),
        gt.get('meno', ''),
        b.get('found_osoby', '')[:80],
        b.get('meno_match', '')
    ]
    
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=row, column=col, value=v)
        c.border = thin
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if col in (5, 8, 11):
            if v == 'Y':
                c.fill = green
            elif v == 'N':
                c.fill = red
        if has_name:
            if col == 1:
                c.font = Font(bold=True)

    row += 1

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 28
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 30
ws2.column_dimensions['H'].width = 10
ws2.column_dimensions['I'].width = 25
ws2.column_dimensions['J'].width = 50
ws2.column_dimensions['K'].width = 10

out = 'benchmark_porovnanie.xlsx'
wb.save(out)
print('Hotovo: ' + out)

# Quick summary
y_meno = sum(1 for u,g in gt_rows.items() if g.get('meno','').strip() and bench_rows.get(u, bench_rows.get(u+'/',{})).get('meno_match')=='Y')
n_meno = sum(1 for u,g in gt_rows.items() if g.get('meno','').strip() and bench_rows.get(u, bench_rows.get(u+'/',{})).get('meno_match')=='N')
print('Meno: ' + str(y_meno) + ' OK, ' + str(n_meno) + ' MISS z 23')
